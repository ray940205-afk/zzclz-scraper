#!/usr/bin/env python3
"""
漳州菜篮子 GitHub Actions 云端抓取脚本
每天自动运行，更新 Excel 文件并提交回仓库
"""

import re
import urllib.request
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ==================== 配置 ====================
OUTPUT_FILE = Path("漳州菜篮子_零售价格日报表.xlsx")
TARGET_URL = "https://zzclz.com/quotations.php"
USER_AGENT = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
CATEGORIES = ['蔬菜', '水果', '肉禽蛋', '水产', '粮油', '其它']

# ==================== 样式定义 ====================
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4DAB3B', end_color='4DAB3B', fill_type='solid')
CAT_FILL = PatternFill(start_color='E8F5E0', end_color='E8F5E0', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
DATA_FONT = Font(name='Arial', size=11)
CAT_FONT = Font(name='Arial', size=11, bold=True, color='2E7D32')
RED_FONT = Font(name='Arial', size=11, bold=True, color='FF0000')
GREEN_FONT = Font(name='Arial', size=11, bold=True, color='008000')
GRAY_FONT = Font(name='Arial', size=11, color='999999')
NOTE_FONT = Font(name='Arial', size=9, color='999999')
STAT_FONT = Font(name='Arial', size=10, bold=True, color='333333')
EVEN_FILL = PatternFill(start_color='F5FAF5', end_color='F5FAF5', fill_type='solid')


def fetch_category_data(date_str, category):
    """抓取指定日期和分类的价格数据"""
    cat_gbk = category.encode('gbk')
    cat_url = urllib.parse.quote(cat_gbk, safe='')
    url = f"{TARGET_URL}?classify={cat_url}&sort={date_str}"

    req = urllib.request.Request(url, headers={
        'User-Agent': USER_AGENT,
        'Accept': 'text/html,application/xhtml+xml,*/*',
        'Referer': TARGET_URL,
    })

    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read()
        html = raw.decode('gbk', errors='replace')
    except Exception as e:
        print(f"    ❌ {category} 请求失败: {e}")
        return None

    pattern = r'<td>&nbsp;&nbsp;(.+?)</td>\s*<td align="center" class="font_Red12px">([\d.]*)\s*</td>.*?<td align="center">(\d{4}-\d{2}-\d{2})</td>'
    matches = re.findall(pattern, html, re.DOTALL)

    result = {}
    for name, price, date in matches:
        name = name.strip()
        if not name or '</td>' in name or '<td' in name:
            continue
        if date == date_str and price:
            result[name] = float(price)
    return result


def fetch_all_categories(date_str):
    """抓取全部 6 个分类"""
    result = {}
    for cat in CATEGORIES:
        data = fetch_category_data(date_str, cat)
        if data:
            result[cat] = data
    return result


def write_sheet(ws, date_str, all_data, prev_all_data=None):
    """写入一个 Sheet"""
    date_obj = datetime.strptime(date_str, '%Y-%m-%d')
    date_display = date_obj.strftime('%Y年%m月%d日')
    weekdays = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']
    weekday = weekdays[date_obj.weekday()]

    ws.merge_cells('A1:E1')
    ws['A1'] = '漳州菜篮子 — 零售价格日报表'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:E2')
    ws['A2'] = f'日期：{date_display}（{weekday}） | 单位：元/500g | 数据来源：漳州菜篮子网'
    ws['A2'].font = Font(name='Arial', size=9, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    headers = ['分类', '品名', '平均价（元/500g）', '较上日涨跌幅', '日期']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN; cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 28

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    row = 4
    total_products = up_count = down_count = unchanged_count = no_data_count = 0

    for category in CATEGORIES:
        data = all_data.get(category, {})
        if not data:
            continue
        prev_data = prev_all_data.get(category, {}) if prev_all_data else {}

        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1, value=f'📂 {category}（{len(data)} 个品种）')
        c.font = CAT_FONT; c.fill = CAT_FILL; c.alignment = LEFT_ALIGN
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).fill = CAT_FILL
        ws.row_dimensions[row].height = 24
        row += 1

        for name, price in data.items():
            yp = prev_data.get(name) if prev_data else None
            total_products += 1

            c = ws.cell(row=row, column=1, value=category)
            c.font = DATA_FONT; c.alignment = CENTER_ALIGN; c.border = THIN_BORDER
            c = ws.cell(row=row, column=2, value=name)
            c.font = DATA_FONT; c.alignment = LEFT_ALIGN; c.border = THIN_BORDER
            c = ws.cell(row=row, column=3, value=price)
            c.font = DATA_FONT; c.alignment = CENTER_ALIGN; c.border = THIN_BORDER
            c.number_format = '0.00'

            if yp is not None and yp > 0:
                change_pct = round(((price - yp) / yp) * 100, 2)
                if change_pct > 0:
                    c = ws.cell(row=row, column=4, value=f'↑ {change_pct:+.2f}%')
                    c.font = RED_FONT; up_count += 1
                elif change_pct < 0:
                    c = ws.cell(row=row, column=4, value=f'↓ {change_pct:+.2f}%')
                    c.font = GREEN_FONT; down_count += 1
                else:
                    c = ws.cell(row=row, column=4, value='→ 0.00%')
                    c.font = GRAY_FONT; unchanged_count += 1
            else:
                c = ws.cell(row=row, column=4, value='—')
                c.font = GRAY_FONT; no_data_count += 1
            c.alignment = CENTER_ALIGN; c.border = THIN_BORDER

            c = ws.cell(row=row, column=5, value=date_str)
            c.font = DATA_FONT; c.alignment = CENTER_ALIGN; c.border = THIN_BORDER

            if row % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = EVEN_FILL
            row += 1

    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1,
            value=f'📊 共 {total_products} 个品种 | 上涨 {up_count} | 下跌 {down_count} | 持平 {unchanged_count} | 无对比 {no_data_count}').font = STAT_FONT
    row += 1
    ws.merge_cells(f'A{row}:E{row}')
    ws.cell(row=row, column=1,
            value='↑ 红色=涨  ↓ 绿色=跌  → 灰色=持平  — =无上日数据').font = NOTE_FONT

    ws.freeze_panes = 'A4'


def main():
    today = datetime.now()
    today_str = today.strftime('%Y-%m-%d')
    sheet_name = today.strftime('%m-%d')

    print(f"\n{'='*55}")
    print(f"  漳州菜篮子 - GitHub Actions 云端抓取")
    print(f"  日期: {today_str}")
    print(f"{'='*55}\n")

    # 1. 抓取今日数据
    print("📡 抓取今日数据...")
    today_all = fetch_all_categories(today_str)
    total_today = sum(len(d) for d in today_all.values())

    if total_today == 0:
        print("  ⚠️  今日暂无数据，跳过")
        return

    for cat in CATEGORIES:
        count = len(today_all.get(cat, {}))
        print(f"  {'✅' if count > 0 else '⚠️'} {cat}: {count} 个")
    print(f"  📊 共计 {total_today} 个品种")

    # 2. 抓取昨日数据
    yesterday = today - timedelta(days=1)
    yesterday_str = yesterday.strftime('%Y-%m-%d')
    print(f"\n📡 抓取昨日数据 ({yesterday_str})...")
    yesterday_all = fetch_all_categories(yesterday_str)
    total_yesterday = sum(len(d) for d in yesterday_all.values())
    print(f"  {'✅' if total_yesterday > 0 else '⚠️'} 共计 {total_yesterday} 个品种")

    # 3. 加载或创建工作簿
    if OUTPUT_FILE.exists():
        print(f"\n📂 加载现有文件")
        wb = load_workbook(OUTPUT_FILE)
    else:
        print(f"\n📂 创建新文件")
        wb = Workbook()
        wb.remove(wb.active)

    # 4. 添加/更新今日 Sheet
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    # 清理空 Sheet
    for s in list(wb.sheetnames):
        if s != sheet_name:
            ws_check = wb[s]
            if ws_check.max_row == 1 and ws_check.max_column == 1 and ws_check['A1'].value is None:
                del wb[s]

    write_sheet(ws, today_str, today_all, yesterday_all if total_yesterday > 0 else None)

    # 5. 按日期排序
    sorted_names = sorted(
        [s for s in wb.sheetnames],
        key=lambda x: datetime.strptime(f"2026-{x}", '%Y-%m-%d'),
        reverse=True
    )
    for i, name in enumerate(sorted_names):
        current_idx = wb.sheetnames.index(name)
        if current_idx != i:
            wb.move_sheet(name, offset=i - current_idx)

    # 6. 保存
    wb.save(OUTPUT_FILE)
    print(f"\n✅ 已保存: {OUTPUT_FILE}")
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   品种: {total_today} 个")


if __name__ == '__main__':
    main()
